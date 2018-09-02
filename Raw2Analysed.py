#Yuman Hordijk 2018

import tkinter as tk
import tkinter.filedialog
import tkinter.tix as tkx
import fnmatch as fn
import glob
from os import chdir, remove, getcwd
from os.path import split
import os
import openpyxl as xl
from openpyxl.utils import *
from openpyxl.chart import *
from openpyxl.drawing import *
from openpyxl.styles import *
from math import log
import statistics
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, Font, RegularTextRun

#Fitted values provided by Thierry Slot for 20% EtOH
##defaultFitValues=["lnA-lnB*expT1*log((flowRate-expY0)/expA1)", 13.55815, 9.21528, 30.82462, -28.99043, 1]
#Fitted values for hexadecane
defaultFitValues=["lnA-lnB*expT1*log((flowRate-expY0)/expA1)", 0.25326, 6.0239, 40.00738, -39.93092, 14.97745]

#Defining of variables
totalVol=0.0
rootPadY=2
rootPadX=2
currFile=0
newFileName=''
headers=['Index', 'Time (s)', 'Bubsize', 'Delta Time (s)', 'Flow rate(bubbles/s)', 'Volume per bubble (uL)', 'Total volume (uL)']
headers2=['File index', 'Number of bubbles', 'Average bubsize', 'Standard deviation', 'Relative 95% CI', 'Total time (s)', 'Total volume (mL)', 'File']
columnWidths=[8, 9, 9, 14, 20, 22, 17]
columnsPerFile=len(headers)+1
files=[]
cutOffValues=[]
topRowOffset=15
leftColumnOffset=1
skipLines=5

lowerBubSizeBound=20
upperBubSizeBound=500

root=tk.Tk()
def SetDefaultFitValues():
    global formula
    formula=defaultFitValues[0]
    global lnB
    lnB=defaultFitValues[1]
    global lnA
    lnA=defaultFitValues[2]
    global expY0
    expY0=defaultFitValues[3]
    global expA1
    expA1=defaultFitValues[4]
    global expT1
    expT1=defaultFitValues[5]

    try:
        entry_formula.delete(0, 'end')
        entry_slope.delete(0, 'end')
        entry_intercept.delete(0, 'end')
        entry_y0.delete(0, 'end')
        entry_A1.delete(0, 'end')
        entry_t1.delete(0, 'end')
        
        entry_formula.insert(0, formula)
        entry_slope.insert(0, lnB)
        entry_intercept.insert(0, lnA)
        entry_y0.insert(0, expY0)
        entry_A1.insert(0, expA1)
        entry_t1.insert(0, expT1)
    except:
        pass
    
def SetFitValues():
    global formula
    formula=entry_formula.get()
    global lnB
    lnB=float(entry_slope.get())
    global lnA
    lnA=float(entry_intercept.get())
    global expY0
    expY0=float(entry_y0.get())
    global expA1
    expA1=float(entry_A1.get())
    global expT1
    expT1=float(entry_t1.get())

def CloseDialog():
    valueDialog.destroy()

def CreateSetFitValuesGUI():
    #Create new GUI
    global valueDialog
    valueDialog=tk.Tk()
    #Formula
    tk.Label(valueDialog, text='Volume(flowRate)= ').grid(row=0, column=0)
    global entry_formula
    entry_formula=tk.Entry(valueDialog)
    entry_formula.grid(row=0, column=1)
    entry_formula.insert(0, formula)
    #slope
    tk.Label(valueDialog, text='Linear Slope (lnB): ').grid(row=1, column=0)
    global entry_slope
    entry_slope=tk.Entry(valueDialog)
    entry_slope.grid(row=1, column=1)
    entry_slope.insert(0, lnB)
    #intercept
    tk.Label(valueDialog, text='Linear Intercept (lnA): ').grid(row=2, column=0)
    global entry_intercept
    entry_intercept=tk.Entry(valueDialog)
    entry_intercept.grid(row=2, column=1)
    entry_intercept.insert(0, lnA)
    #y0
    tk.Label(valueDialog, text='Exp. y0 (expY0): ').grid(row=3, column=0)
    global entry_y0
    entry_y0=tk.Entry(valueDialog)
    entry_y0.grid(row=3, column=1)
    entry_y0.insert(0, expY0)
    #A1
    tk.Label(valueDialog, text='Exp. A1 (expA1): ').grid(row=4, column=0)
    global entry_A1
    entry_A1=tk.Entry(valueDialog)
    entry_A1.grid(row=4, column=1)
    entry_A1.insert(0, expA1)
    #t1
    tk.Label(valueDialog, text='Exp. t1 (expT1): ').grid(row=5, column=0)
    global entry_t1
    entry_t1=tk.Entry(valueDialog)
    entry_t1.grid(row=5, column=1)
    entry_t1.insert(0, expT1)
    #Button to close GUI and set the values
    tk.Button(valueDialog, text='Set parameters', command=SetFitValues).grid(row=6, column=0)
    tk.Button(valueDialog, text='Reset to default', command=SetDefaultFitValues).grid(row=6, column=1)
    tk.Button(valueDialog, text='Close', command=CloseDialog).grid(row=6, column=2)
    valueDialog.mainloop()

def GetFiles():
    #Ask file directory and select all *_bc_ files
    global cutOffValues
    global maxRows
    global files

    files=tk.filedialog.askopenfilenames(filetypes=(("BC files","*bc*"),))

    #Create empty (0) cutoffvalues and maxRows list
    cutOffValues=[]
    maxRows=[]
    for i in files:
        with open(i, 'r') as f:
            cutOffValues.append(0)
            maxRows.append(len(f.readlines())-skipLines)

    chdir(split(files[0])[0])
    print('Working directory: %s' % getcwd())

def CreateHeaders(ws, file, fileIndex):
    ws.cell(row=topRowOffset, column=fileIndex*columnsPerFile+leftColumnOffset).value=file

    for p, i in enumerate(headers):
        ws.cell(row=topRowOffset+1, column=fileIndex*columnsPerFile+leftColumnOffset+p).value=i
        ws.column_dimensions[get_column_letter(fileIndex*columnsPerFile+p+1)].width=columnWidths[p]

def CreateCharts(ws, fileIndex):
    chartObj=ScatterChart()

    min_col=fileIndex*columnsPerFile+leftColumnOffset
    min_row=topRowOffset+2
    max_col=fileIndex*columnsPerFile+leftColumnOffset
    max_row=maxRows[fileIndex]-cutOffValues[fileIndex]+topRowOffset+1
    
    xValues=Reference(ws, min_col=min_col+headers.index('Time (s)'), min_row=min_row, max_col=max_col+headers.index('Time (s)'), max_row=max_row)
    yValues=Reference(ws, min_col=min_col+headers.index('Total volume (uL)'), min_row=min_row, max_col=max_col+headers.index('Total volume (uL)'), max_row=max_row)
    
    serObj=Series(yValues, xValues, title='Volume vs. time')
    chartObj.append(serObj)
    chartObj.x_axis.title='Time (s)'
    chartObj.y_axis.title='Volume (uL)'
    chartObj.legend=None

    #<source=DrFunn1 at https://python-forum.io/Thread-How-to-change-font-size-of-chart-title-and-axis-title>
    font = Font(typeface='Verdana')
    size = 1000
    cp = CharacterProperties(latin=font, sz=size, b=False)
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    chartObj.x_axis.txPr = rtp
    chartObj.y_axis.txPr = rtp

    chartObj.x_axis.title.tx.rich.p[0].pPr = pp
    chartObj.y_axis.title.tx.rich.p[0].pPr = pp
    #</source>

    ws.add_chart(chartObj, get_column_letter(fileIndex*columnsPerFile+leftColumnOffset)+'1')

def SetCutOffValues():
    global cutOffValues
    for i in range(0, len(cutOffValues)):
        try:
            if int(fileEntries[i].get())>maxRows[i]:
                cutOffValues[i]=int(maxRows[i])
            else:
                cutOffValues[i]=int(fileEntries[i].get())
        except:
            pass
    fileListDialog.destroy()

def ShowFileList():
    #Create new GUI
    global fileListDialog
    global fileEntries
    global cutOffValues
    global maxRows
    fileEntries=[]
    fileListDialog=tk.Tk()

    #Check if files have been selected
    if files==[]:
        tk.Label(fileListDialog, text='No files found! First select directory.').pack()
    #If it has, than build GUI
    else:
        for row, fileName in enumerate(files):

            tk.Label(fileListDialog, text='#'+str(row)).grid(row=row, column=0, pady=rootPadY, padx=rootPadX)
            tk.Label(fileListDialog, text='Left cut-off: ').grid(row=row, column=1, pady=rootPadY, padx=rootPadX)
            fileEntries.append(tk.Entry(fileListDialog))
            fileEntries[row].grid(row=row, column=2, pady=rootPadY, padx=rootPadX)
            fileEntries[row].insert(0, cutOffValues[row])
            cutOffValues[row]=int(fileEntries[row].get())
            tk.Label(fileListDialog, text='Max rows: '+str(maxRows[row])).grid(row=row, column=3, pady=rootPadY, padx=rootPadX)
            tk.Label(fileListDialog, text=str(fileName), anchor='w').grid(sticky='w', row=row, column=4, pady=rootPadY, padx=rootPadX)

        tk.Button(fileListDialog, text='Set cut-offs', command=SetCutOffValues).grid(row=row+1, columnspan=2)

    fileListDialog.mainloop()

def AddDataToFile(ws, file, fileIndex, currRow, data, mode):
    try:
        newValue=0.0
        if mode==0:
            #Add index for easy cutting
            newValue=currRow+cutOffValues[fileIndex]
            
        if mode==1:
            #Add time data
            newValue=float(data.split(',')[0])-float(cutOffTime)
            
        if mode==2:
            #Add bubsize data
            newValue=float(data.split(',')[1])
            
        if mode==3:
            #Add delta time
            if currRow>0:
                newValue=float(ws.cell(row=currRow+topRowOffset+2, column=fileIndex*columnsPerFile+leftColumnOffset+headers.index('Time (s)')).value)-float(ws.cell(row=currRow+topRowOffset+1, column=fileIndex*columnsPerFile+leftColumnOffset+headers.index('Time (s)')).value)
                if currRow==1:
                    #Copy second point to first point.
                    ws.cell(row=topRowOffset+2, column=fileIndex*columnsPerFile+leftColumnOffset+mode).value=newValue
                    
        if mode==4:
            #Calculate flow rate
                newValue=1.0/float(ws.cell(row=currRow+topRowOffset+2, column=fileIndex*columnsPerFile+leftColumnOffset+headers.index('Delta Time (s)')).value)

        if mode==5:
            try:
                flowRate=ws.cell(row=currRow+topRowOffset+2, column=mode+fileIndex*columnsPerFile+leftColumnOffset-1).value
                newValue=eval(formula)
                
            except:
                try:
                    print("Error: could not evaluate expression at line %d, flowRate=%d") %(currRow, flowRate)
                except:
                    pass


        if mode==6:
            #Sum volume
            global totalVol
            totalVol+=float(ws.cell(row=currRow+topRowOffset+2, column=mode+fileIndex*columnsPerFile+leftColumnOffset-1).value)
            newValue=totalVol

        ws.cell(row=currRow+topRowOffset+2, column=mode+fileIndex*columnsPerFile+leftColumnOffset).value=newValue

    except:
        print('Error when adding data to file, CurrRow: %d, Mode: %d, Data: %s, Position: %s' %(currRow, mode, data.strip('\n'), get_column_letter(fileIndex*columnsPerFile+leftColumnOffset+mode)+str(currRow+topRowOffset+2)))
        pass

def GetBubSizes(index):
    bubSizes=[]
    for a in range(0, maxRows[index]):
        v=ws1.cell(row=topRowOffset+2+a, column=index*columnsPerFile+leftColumnOffset+headers.index('Bubsize')).value
        bubSizes.append(v)
    return bubSizes

def AddHeaders():
    for a, header in enumerate(headers2):
        ws2.cell(row=1, column=a+1).value=header

def AddVariousData(index, fileName, bubSizes):
    #headers2=['File index', 'Number of bubbles', 'Average bubsize', 'Standard deviation', 'Relative 95% CI', 'Total time (s)', 'Total volume (mL)', 'File']   
    for a, header in enumerate(headers2):
        try:
            if a==0:
                newValue=index
            if a==1:
                newValue=maxRows[index]-cutOffValues[index]
            if a==2:
                col=get_column_letter(index*columnsPerFile+leftColumnOffset+headers.index('Bubsize'))
                newValue="=Average('%s'!%s%d:%s%d)" %(ws1.title, col, topRowOffset+2, col, topRowOffset+2+maxRows[index])
            if a==3:
                col=get_column_letter(index*columnsPerFile+leftColumnOffset+headers.index('Bubsize'))
                newValue="=_xlfn.STDEV.S('%s'!%s%d:%s%d)" %(ws1.title, col, topRowOffset+2, col, topRowOffset+2+maxRows[index])
            if a==4:
##                stdev=ws2.cell(row=2+index, column=headers2.index('Standard deviation')).value
##                count=ws2.cell(row=2+index, column=headers2.index('Number of bubbles')).value
##                newValue="=CONFIDENCE(0.05;%d;%d)" %(stdev, count)
                pass
            if a==5:
##              newValue=float(totalTime)
                pass
            if a==6:
                newValue=totalVol/1000
            if a==7:
                newValue=fileName

            ws2.cell(row=2+index, column=a+1).value=newValue
        except:
            pass
        
def Main():
    #Create new .xlsx file
    wb=xl.Workbook()
    wb.save(entry_a.get()+'.xlsx')
    #Load new .xlsx file and create and load sheets
    wb=xl.load_workbook(entry_a.get()+'.xlsx')
    global ws1
    ws1=wb.active
    ws1.title='Raw Data'
    global ws2
    ws2=wb.create_sheet('Various data')

    AddHeaders()
    
    #Loop over all the files
    for a in range(0, len(headers)):
        rejectedValues=0
        for b, fileName in enumerate(files):
            CreateHeaders(ws1, fileName, b)
            #open file and create a temp file
            global totalVol
            totalVol=0.0            
            with open(fileName, 'r') as file, open('temp', 'a+') as temp:
                global cutOffTime
                #Get lines and filter them for the right format
                lines=fn.filter(file.readlines(), '*, *')
                #Write all usable lines to temp file
                for c, line in enumerate(lines):
                    #skip first 4 rows containing non-usable data and skip to specified cutoffvalue
                    if c==cutOffValues[b]:
                        cutOffTime=line.split(',')[0]
                    if c>=cutOffValues[b]:
                        #Check if bubsize is between specified bounds
                        try:
                            #Check if bubble size is within bounds
                            if float(line.split(',')[1])>=lowerBubSizeBound and float(line.split(',')[1])<=upperBubSizeBound:
                                temp.write(line)
                            else:
                                rejectedValues+=1
                        except:
                            print(line, c, type(line))
                            raise

            with open('temp', 'r') as temp:
                #Loop over all lines in temp
                print("Stage "+str(b+1)+'/'+str(len(files)), "Rejected "+str(rejectedValues)+" bubbles" , fileName)
                #Get lines from temp
                lines=temp.readlines()
                #get total time
##                global totalTime
##                totalTime=lines[-1].split(',')[0]
                for c, line in enumerate(lines):
                    #Select time data and add to file
                    AddDataToFile(ws1, fileName, b, c, line, a)

            #Add various data to sheet ws2
            if a==len(headers)-1:
                bubSizes=GetBubSizes(b)
                AddVariousData(b, fileName, bubSizes)
            #Remove temp file
            remove('temp')

    for a in range(len(files)):
        CreateCharts(ws1, a)

    wb.save(entry_a.get()+'.xlsx')
    os.system('start excel.exe "%s"' % (entry_a.get()+'.xlsx'))

def CreateGUI():
    #Base GUI
    tk.Label(root, text='Enter new file name:').grid(row=2, column=0, pady=rootPadY, padx=rootPadX)
    global entry_a
    entry_a=tk.Entry(root)
    entry_a.grid(sticky='w', row=2, column=1, columnspan=5)
    frame=tk.Frame(root)
    frame.grid(row=3, columnspan=2)
    tk.Button(frame, text='Select files', command=GetFiles).grid(row=1, column=0, pady=rootPadY, padx=rootPadX)
    tk.Button(frame, text='Start program', command=Main).grid(row=1, column=1, pady=rootPadY, padx=rootPadX)
    tk.Button(frame, text='Set fitted Parameters', command=CreateSetFitValuesGUI).grid(row=1, column=2, pady=rootPadY, padx=rootPadX)
    tk.Button(frame, text='Show file list', command=ShowFileList).grid(row=1, column=3, pady=rootPadY, padx=rootPadX)

SetDefaultFitValues()
CreateGUI()
root.mainloop()
